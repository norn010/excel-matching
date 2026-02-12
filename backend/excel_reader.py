# -*- coding: utf-8 -*-
"""
Utilities สำหรับอ่านข้อมูล Excel (.xlsx / .xls)
และตรวจจับลำดับคอลัมน์จากชื่อหัวตารางอัตโนมัติ
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None

try:
    import xlrd
except Exception:  # pragma: no cover
    xlrd = None


COMPARE_COLUMN_LABELS = ["ชื่อลูกค้า", "ไฟแนนซ์", "Moder Code", "เลขถัง", "ราคาขาย", "COM F/N", "COM"]
COMPARE_FIELDS = ["customer", "finance", "model", "tank", "price", "com_fn", "com"]

COMPARE_FIELD_ALIASES = {
    "customer": ["ชื่อลูกค้า", "ลูกค้า", "customer", "name"],
    "finance": ["ไฟแนนซ์", "ไฟแนน", "finance"],
    "model": ["moder code", "model code", "model", "รุ่น"],
    "tank": ["เลขถัง", "vin", "เลข vin", "ตัวถัง"],
    # ใช้ "ราคาขาย" เป็นหลัก (ไม่ใช้ Wholesales)
    "price": ["ราคาขาย", "sale price", "selling price"],
    "com_fn": ["com f/n", "com fn", "comf/n"],
    "com": ["com"],
}


def _norm_text(v: object) -> str:
    s = str(v or "").strip().lower()
    s = s.replace("\n", " ")
    s = re.sub(r"\s+", " ", s)
    return s


def _cell_to_str(v: object) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return s


def _is_vin_like(s: str) -> bool:
    t = re.sub(r"[^a-zA-Z0-9]", "", s or "")
    return len(t) >= 10 and bool(re.search(r"[a-zA-Z]", t)) and bool(re.search(r"\d", t))


def _norm_sheet_name(s: str) -> str:
    return re.sub(r"\s+", "", (s or "")).lower()


def _score_sheet_matrix(matrix: list[list[str]]) -> int:
    score = 0
    for row in matrix[:10]:
        text = " ".join(_norm_text(x) for x in row)
        if "moder code" in text:
            score += 5
        if "com f/n" in text:
            score += 5
        if "vin" in text or "เลขถัง" in text:
            score += 3
    return score


def _load_matrix_xlsx(path: str, sheet_name: Optional[str], max_rows: int = 40, max_cols: int = 60):
    if openpyxl is None:
        raise ValueError("ต้องติดตั้ง openpyxl เพื่ออ่านไฟล์ .xlsx")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        selected = None
        if sheet_name and sheet_name in wb.sheetnames:
            selected = sheet_name
        if selected is None:
            best_score = -1
            for sn in wb.sheetnames:
                ws = wb[sn]
                matrix = []
                for r in range(1, max_rows + 1):
                    row = []
                    for c in range(1, max_cols + 1):
                        row.append(_cell_to_str(ws.cell(r, c).value))
                    matrix.append(row)
                sc = _score_sheet_matrix(matrix)
                if sc > best_score:
                    best_score = sc
                    selected = sn
        ws = wb[selected or wb.sheetnames[0]]
        matrix = []
        for r in range(1, max_rows + 1):
            row = []
            for c in range(1, max_cols + 1):
                row.append(_cell_to_str(ws.cell(r, c).value))
            matrix.append(row)
        return matrix, ws.title
    finally:
        wb.close()


def _load_matrix_xls(path: str, sheet_name: Optional[str], max_rows: int = 40, max_cols: int = 60):
    if xlrd is None:
        raise ValueError("ต้องติดตั้ง xlrd (1.x) เพื่ออ่านไฟล์ .xls")
    wb = xlrd.open_workbook(path)
    selected = None
    if sheet_name:
        try:
            selected = wb.sheet_by_name(sheet_name)
        except Exception:
            selected = None
    if selected is None:
        best_score = -1
        for i in range(wb.nsheets):
            sh = wb.sheet_by_index(i)
            matrix = []
            for r in range(min(max_rows, sh.nrows)):
                row = []
                for c in range(min(max_cols, sh.ncols)):
                    row.append(_cell_to_str(sh.cell_value(r, c)))
                matrix.append(row)
            sc = _score_sheet_matrix(matrix)
            if sc > best_score:
                best_score = sc
                selected = sh
    sh = selected if selected is not None else wb.sheet_by_index(0)
    matrix = []
    for r in range(min(max_rows, sh.nrows)):
        row = []
        for c in range(min(max_cols, sh.ncols)):
            row.append(_cell_to_str(sh.cell_value(r, c)))
        matrix.append(row)
    return matrix, sh.name


def _detect_cols_and_start(matrix: list[list[str]], fallback_cols: list[int]):
    max_cols = max((len(r) for r in matrix), default=0)
    col_blobs: list[str] = []
    for c in range(max_cols):
        parts = []
        for r in range(min(12, len(matrix))):
            v = matrix[r][c] if c < len(matrix[r]) else ""
            nv = _norm_text(v)
            if nv:
                parts.append(nv)
        col_blobs.append(" ".join(parts))

    chosen: dict[str, int] = {}
    scores: dict[str, int] = {}
    used: set[int] = set()
    for fi, field in enumerate(COMPARE_FIELDS):
        aliases = COMPARE_FIELD_ALIASES[field]
        best_col = None
        best_score = 0
        for c, blob in enumerate(col_blobs):
            if c in used:
                continue
            score = 0
            for a in aliases:
                if a in blob:
                    score = max(score, 10 + len(a))
            if field == "com" and "com f/n" in blob:
                score -= 8
            if score > best_score:
                best_score = score
                best_col = c
        if best_col is None or best_score <= 0:
            cand = fallback_cols[fi]
            if cand in used:
                for x in range(max_cols):
                    if x not in used:
                        cand = x
                        break
            best_col = cand
        chosen[field] = best_col
        scores[field] = best_score
        used.add(best_col)

    tank_col = chosen["tank"]
    customer_col = chosen["customer"]
    start_row = 6
    for r in range(len(matrix)):
        tank = _norm_text(matrix[r][tank_col] if tank_col < len(matrix[r]) else "")
        cust = _norm_text(matrix[r][customer_col] if customer_col < len(matrix[r]) else "")
        if _is_vin_like(tank) and cust and "ลูกค้า" not in cust and "ลำดับ" not in cust:
            start_row = r
            break

    cols = [chosen[k] for k in COMPARE_FIELDS]
    return cols, start_row, scores


def detect_compare_layout(file_path: str, sheet_name: Optional[str], fallback_cols: list[int]):
    p = Path(file_path).resolve()
    suf = p.suffix.lower()
    if suf == ".xls":
        matrix, detected_sheet = _load_matrix_xls(str(p), sheet_name)
    elif suf in (".xlsx", ".xlsm"):
        matrix, detected_sheet = _load_matrix_xlsx(str(p), sheet_name)
    else:
        raise ValueError(f"ไม่รองรับนามสกุลไฟล์: {suf}")
    cols, start_row, scores = _detect_cols_and_start(matrix, fallback_cols=fallback_cols)
    return cols, start_row, detected_sheet, scores


def read_esg_rows(
    file_path: str,
    start_row: int = 6,
    col_indices: Optional[list[int]] = None,
    sheet_name: Optional[str] = None,
) -> list[list[str]]:
    if col_indices is None:
        col_indices = [2, 3, 5, 6, 8, 14, 15]
    p = Path(file_path).resolve()
    suf = p.suffix.lower()
    if suf in (".xlsx", ".xlsm"):
        if openpyxl is None:
            raise ValueError("ต้องติดตั้ง openpyxl เพื่ออ่านไฟล์ .xlsx")
        wb = openpyxl.load_workbook(str(p), data_only=True, read_only=True)
        try:
            ws = None
            if sheet_name:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    want = _norm_sheet_name(sheet_name)
                    for sn in wb.sheetnames:
                        cur = _norm_sheet_name(sn)
                        if want and (want in cur or cur in want):
                            ws = wb[sn]
                            break
            if ws is None:
                # fallback หา sheet งานเงินรางวัล ESG
                for sn in wb.sheetnames:
                    cur = _norm_sheet_name(sn)
                    if ("esg" in cur) and ("รางวัล" in sn or "เรียกเก็บ" in sn):
                        ws = wb[sn]
                        break
            if ws is None:
                ws = wb[wb.sheetnames[0]]
            rows_out = []
            for r in range(start_row + 1, ws.max_row + 1):
                vals = []
                for c in col_indices:
                    vals.append(_cell_to_str(ws.cell(r, c + 1).value))
                if any(vals):
                    rows_out.append(vals)
            return rows_out
        finally:
            wb.close()
    if suf == ".xls":
        if xlrd is None:
            raise ValueError("ต้องติดตั้ง xlrd (1.x) เพื่ออ่านไฟล์ .xls")
        wb = xlrd.open_workbook(str(p))
        sh = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)
        rows_out = []
        for r in range(start_row, sh.nrows):
            vals = []
            for c in col_indices:
                vals.append(_cell_to_str(sh.cell_value(r, c)) if c < sh.ncols else "")
            if any(vals):
                rows_out.append(vals)
        return rows_out
    raise ValueError(f"ไม่รองรับนามสกุลไฟล์: {suf}")


def read_tax_sheet_rows(
    file_path: str,
    sheet_name: str = "ตารางไมวัน",
    col_indices: Optional[list[int]] = None,
    start_row: int = 0,
) -> list[list[str]]:
    if col_indices is None:
        col_indices = [1, 2, 4, 5, 7, 13, 14]
    p = Path(file_path).resolve()
    suf = p.suffix.lower()
    if suf == ".xls":
        if xlrd is None:
            raise ValueError("ต้องติดตั้ง xlrd (1.x) เพื่ออ่านไฟล์ .xls")
        wb = xlrd.open_workbook(str(p))
        try:
            sh = wb.sheet_by_name(sheet_name)
        except Exception:
            # fallback: จับชื่อชีตแบบใกล้เคียง (เช่น พิมพ์ "ไมวัน" แทน "ตารางไมวัน")
            sh = None
            want = _norm_sheet_name(sheet_name)
            for i in range(wb.nsheets):
                cand = wb.sheet_by_index(i)
                cname = _norm_sheet_name(cand.name)
                if want and (want in cname or cname in want):
                    sh = cand
                    break
            if sh is None:
                sh = wb.sheet_by_index(0)
        rows_out = []
        for r in range(start_row, sh.nrows):
            vals = []
            for c in col_indices:
                vals.append(_cell_to_str(sh.cell_value(r, c)) if c < sh.ncols else "")
            if any(vals):
                rows_out.append(vals)
        return rows_out
    if suf in (".xlsx", ".xlsm"):
        if openpyxl is None:
            raise ValueError("ต้องติดตั้ง openpyxl เพื่ออ่านไฟล์ .xlsx")
        wb = openpyxl.load_workbook(str(p), data_only=True, read_only=True)
        try:
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
            rows_out = []
            for r in range(start_row + 1, ws.max_row + 1):
                vals = []
                for c in col_indices:
                    vals.append(_cell_to_str(ws.cell(r, c + 1).value))
                if any(vals):
                    rows_out.append(vals)
            return rows_out
        finally:
            wb.close()
    raise ValueError(f"ไม่รองรับนามสกุลไฟล์: {suf}")


def get_sheet_names(file_path: str) -> list[str]:
    p = Path(file_path).resolve()
    if not p.exists():
        return []
    suf = p.suffix.lower()
    if suf == ".xls":
        if xlrd is None:
            return []
        wb = xlrd.open_workbook(str(p))
        return [wb.sheet_by_index(i).name for i in range(wb.nsheets)]
    if suf in (".xlsx", ".xlsm"):
        if openpyxl is None:
            return []
        wb = openpyxl.load_workbook(str(p), data_only=True, read_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    return []
